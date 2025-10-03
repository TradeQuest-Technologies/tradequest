"""
Journal endpoints
"""

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Body
from sqlalchemy.orm import Session
from typing import List, Optional
import structlog
import pandas as pd
import io
import json
import math
import uuid
from datetime import datetime

from app.core.database import get_db
from app.core.auth import get_current_user
from app.schemas.trade import TradeCreate, TradeResponse, JournalEntryCreate, JournalEntryResponse
from app.models.user import User
from app.models.trade import Trade, JournalEntry
from app.services.ai_csv_converter import AICSVConverter
from app.services.csv_parser import CSVParser
from app.services.file_upload_service import FileUploadService

logger = structlog.get_logger()

def safe_numeric_value(value):
    """Convert numeric value to JSON-safe format, handling NaN/Infinity"""
    if value is None:
        return 0.0
    
    try:
        # Convert to float to check for special values
        float_val = float(value)
        if math.isnan(float_val) or math.isinf(float_val):
            logger.warning("Found problematic numeric value", value=value, float_val=float_val)
            return 0.0
        return float_val
    except (ValueError, TypeError, OverflowError) as e:
        logger.warning("Failed to convert numeric value", value=value, error=str(e), type=type(value))
        return 0.0
    except Exception as e:
        logger.error("Unexpected error in safe_numeric_value", value=value, error=str(e), type=type(value))
        return 0.0

def safe_raw_value(value):
    """Convert raw field to JSON-safe format"""
    if value is None:
        return None
    
    if isinstance(value, str):
        try:
            # Try to parse as JSON and clean any problematic values
            parsed = json.loads(value)
            logger.info("Successfully parsed raw JSON", raw_length=len(value))
            return json.dumps(parsed)  # Re-serialize to ensure it's clean
        except json.JSONDecodeError as e:
            logger.warning("Failed to parse raw JSON", error=str(e), raw_preview=value[:100])
            return value  # Return as-is if not valid JSON
    logger.info("Raw value is not a string", type=type(value))
    return value

def is_trade_corrupted(trade):
    """Check if a trade has corrupted data that would cause JSON serialization issues"""
    try:
        # Check numeric fields for problematic values
        numeric_fields = ['qty', 'avg_price', 'fees', 'pnl']
        for field in numeric_fields:
            value = getattr(trade, field, None)
            if value is not None:
                try:
                    float_val = float(value)
                    if math.isnan(float_val) or math.isinf(float_val):
                        logger.warning(f"Corrupted trade detected - {field} has invalid value", 
                                     trade_id=trade.id, field=field, value=value)
                        return True, f"{field} has invalid value: {value}"
                except (ValueError, TypeError, OverflowError):
                    logger.warning(f"Corrupted trade detected - {field} cannot be converted to float", 
                                 trade_id=trade.id, field=field, value=value, type=type(value))
                    return True, f"{field} cannot be converted to float: {value}"
        
        # Check raw field for JSON issues
        if trade.raw is not None:
            try:
                if isinstance(trade.raw, str):
                    json.loads(trade.raw)  # Try to parse as JSON
            except json.JSONDecodeError:
                logger.warning(f"Corrupted trade detected - raw field has invalid JSON", 
                             trade_id=trade.id, raw_preview=str(trade.raw)[:100])
                return True, "raw field has invalid JSON"
        
        return False, None
        
    except Exception as e:
        logger.error(f"Error checking trade corruption", trade_id=trade.id, error=str(e))
        return True, f"Error validating trade: {str(e)}"
router = APIRouter()

@router.post("/ingest_csv")
async def ingest_csv(
    file: UploadFile = File(...),
    venue: str = Form(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Import trades from CSV file"""
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be a CSV"
        )
    
    try:
        # Read CSV content
        content = await file.read()
        
        # Try to parse CSV with error handling
        try:
            df = pd.read_csv(io.StringIO(content.decode('utf-8')))
        except pd.errors.ParserError as parse_error:
            # Try parsing with more flexible options
            try:
                df = pd.read_csv(
                    io.StringIO(content.decode('utf-8')),
                    on_bad_lines='skip',  # Skip malformed lines
                    engine='python'       # Use Python engine for better error handling
                )
                logger.warning("CSV parsing had issues, skipped malformed lines", 
                             user_id=str(current_user.id), error=str(parse_error))
            except Exception as e:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"CSV file is malformed and cannot be parsed: {str(e)}"
                )
        
        if df.empty:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="CSV file is empty or contains no valid data"
            )
        
        # Parse CSV using our parser
        parser = CSVParser()
        try:
            trades_data = parser.parse_trades(df)
        except ValueError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"CSV contains corrupted data that cannot be processed: {str(e)}"
            )
        
        # Create trade records, filtering out corrupted ones
        created_trades = []
        corrupted_count = 0
        
        for trade_data in trades_data:
            try:
                # Apply safe numeric conversion BEFORE creating Trade object
                original_qty = trade_data.get('qty')
                original_price = trade_data.get('avg_price')
                original_fees = trade_data.get('fees')
                original_pnl = trade_data.get('pnl')
                
                trade_data['qty'] = safe_numeric_value(original_qty)
                trade_data['avg_price'] = safe_numeric_value(original_price)
                trade_data['fees'] = safe_numeric_value(original_fees)
                trade_data['pnl'] = safe_numeric_value(original_pnl)
                
                # Debug logging
                if any(math.isnan(float(v)) if isinstance(v, (int, float)) and v is not None else False for v in [trade_data['qty'], trade_data['avg_price'], trade_data['fees'], trade_data['pnl']]):
                    logger.error("Found NaN values after safe_numeric_value", 
                               original_qty=original_qty, original_price=original_price, 
                               original_fees=original_fees, original_pnl=original_pnl,
                               processed_qty=trade_data['qty'], processed_price=trade_data['avg_price'],
                               processed_fees=trade_data['fees'], processed_pnl=trade_data['pnl'])
                
                # Convert raw dict to JSON string for database storage
                raw_data = trade_data.get('raw', {})
                if isinstance(raw_data, dict):
                    trade_data['raw'] = json.dumps(raw_data)
                
                trade = Trade(
                    user_id=current_user.id,
                    venue=venue,  # Use the selected venue
                    **trade_data
                )
                
                db.add(trade)
                created_trades.append(trade)
                
            except Exception as e:
                corrupted_count += 1
                logger.warning("Failed to create trade from CSV data", error=str(e), trade_data=trade_data)
                continue
        
        db.commit()
        
        # Refresh to get IDs
        for trade in created_trades:
            db.refresh(trade)
        
        logger.info("CSV ingested", user_id=str(current_user.id), count=len(created_trades))
        
        # Convert trades to response format, handling NaN/Infinity values
        trade_responses = []
        for trade in created_trades:
            try:
                trade_dict = {
                    'id': trade.id,
                    'user_id': trade.user_id,
                    'account': trade.account,
                    'venue': trade.venue,
                    'symbol': trade.symbol,
                    'side': trade.side,
                    'qty': safe_numeric_value(trade.qty),
                    'avg_price': safe_numeric_value(trade.avg_price),
                    'fees': safe_numeric_value(trade.fees),
                    'pnl': safe_numeric_value(trade.pnl),
                    'submitted_at': trade.submitted_at,
                    'filled_at': trade.filled_at,
                    'order_ref': trade.order_ref,
                    'session_id': trade.session_id,
                    'raw': safe_raw_value(trade.raw),
                    'chart_image': trade.chart_image
                }
                trade_responses.append(TradeResponse(**trade_dict))
            except Exception as e:
                logger.warning("Failed to convert trade to response", trade_id=trade.id, error=str(e))
                continue
        
        message = f"Successfully imported {len(created_trades)} trades"
        if corrupted_count > 0:
            message += f" (skipped {corrupted_count} corrupted trades)"
        
        return {
            "message": message,
            "ingested": len(created_trades),
            "skipped_corrupted": corrupted_count,
            "trades": trade_responses
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error("CSV ingestion failed", user_id=str(current_user.id), error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to import CSV: {str(e)}"
        )

@router.post("/ai_convert_csv_preview")
async def ai_convert_csv_preview(
    file: UploadFile = File(...),
    venue: str = Form(...),
    import_type: str = Form("positions"),  # "trades" or "positions"
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """AI-powered CSV conversion - returns preview without saving to database
    
    Args:
        import_type: 
            - "positions": Each row is already a complete position (entry + exit grouped)
            - "trades": Raw individual trades that need to be grouped into position cycles
    """
    
    # Check file type
    allowed_extensions = ['.csv', '.xlsx', '.xls']
    if not any(file.filename.endswith(ext) for ext in allowed_extensions):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be CSV or Excel (.csv, .xlsx, .xls)"
        )
    
    try:
        # Read file content
        content = await file.read()
        
        # Convert Excel to CSV if needed
        if file.filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            from io import BytesIO, StringIO
            import csv as csv_module
            
            # Load Excel file
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
            ws = wb.active
            
            # Convert to CSV string using proper CSV writer (handles commas, quotes, etc.)
            output = StringIO()
            csv_writer = csv_module.writer(output)
            
            for row in ws.iter_rows(values_only=True):
                # Convert row values to strings and handle None
                csv_writer.writerow([str(cell) if cell is not None else '' for cell in row])
            
            csv_content = output.getvalue()
            logger.info("Converted Excel to CSV", 
                       filename=file.filename,
                       rows=ws.max_row)
        else:
            # Already CSV
            csv_content = content.decode('utf-8')
        
        # Use AI to convert the CSV
        ai_converter = AICSVConverter()
        
        # Pass the import type to the AI converter
        group_into_positions = (import_type == "trades")
        result = await ai_converter.convert_csv_with_ai(
            csv_content, 
            file.filename,
            group_positions=group_into_positions
        )
        
        if not result["success"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"AI conversion failed: {result['error']}"
            )
        
        logger.info("CSV conversion completed", 
                   import_type=import_type,
                   group_into_positions=group_into_positions,
                   positions_count=len(result.get("positions", [])))
        
        # Convert positions to preview format (without saving to database)
        preview_trades = []
        
        for position_data in result["positions"]:
            try:
                # Parse datetime strings
                filled_at_str = position_data.get('filled_at')
                submitted_at_str = position_data.get('submitted_at')
                
                filled_at_dt = None
                if filled_at_str:
                    try:
                        filled_at_dt = datetime.fromisoformat(filled_at_str.replace('Z', '+00:00'))
                    except Exception as e:
                        logger.warning("Failed to parse filled_at", value=filled_at_str, error=str(e))
                        filled_at_dt = datetime.utcnow()
                
                submitted_at_dt = None
                if submitted_at_str:
                    try:
                        submitted_at_dt = datetime.fromisoformat(submitted_at_str.replace('Z', '+00:00'))
                    except Exception as e:
                        logger.warning("Failed to parse submitted_at", value=submitted_at_str, error=str(e))
                
                # Create preview trade object
                # Get raw data to extract entry/exit prices
                raw_data = position_data.get('raw', {})
                side = position_data.get('side', 'buy')
                
                # Get the raw entry/exit prices from AI converter
                ai_entry_price = safe_numeric_value(raw_data.get('entry_price', 0))
                ai_exit_price = safe_numeric_value(raw_data.get('exit_price', 0))
                
                # For LONGS: entry = buy price, exit = sell price
                # For SHORTS: entry = sell price (opening short), exit = buy price (closing short)
                if side == 'long':
                    entry_price = ai_entry_price  # This is the buy price
                    exit_price = ai_exit_price    # This is the sell price
                else:  # short
                    entry_price = ai_exit_price   # Shorts open with a sell
                    exit_price = ai_entry_price   # Shorts close with a buy
                
                # For display, use entry price as the main price
                display_price = entry_price if entry_price > 0 else exit_price
                
                preview_trade = {
                    'id': str(uuid.uuid4()),  # Temporary ID for frontend tracking
                    'symbol': position_data.get('symbol', ''),
                    'side': side,
                    'qty': safe_numeric_value(position_data.get('qty', 0)),
                    'avg_price': display_price,
                    'entry_price': entry_price,
                    'exit_price': exit_price,
                    'fees': safe_numeric_value(position_data.get('fees', 0)),
                    'pnl': safe_numeric_value(position_data.get('pnl', 0)),
                    'filled_at': filled_at_dt.isoformat() if filled_at_dt else None,
                    'submitted_at': submitted_at_dt.isoformat() if submitted_at_dt else None,
                    'venue': venue,
                    'account': 'spot',
                    'raw': raw_data
                }
                
                preview_trades.append(preview_trade)
                
            except Exception as e:
                logger.warning("Failed to create preview trade", error=str(e), position_data=position_data)
        
        logger.info("AI CSV preview generated", 
                   user_id=str(current_user.id), 
                   positions_count=len(preview_trades),
                   broker_detected=result.get("broker_detected", "Unknown"))
        
        return {
            "message": f"AI analyzed {len(preview_trades)} positions from your CSV",
            "trades": preview_trades,
            "total_positions": len(preview_trades),
            "broker_detected": result.get("broker_detected", "Unknown"),
            "conversion_mapping": result.get("mapping", {}),
            "original_rows": result.get("original_rows", 0),
            "venue": venue
        }
        
    except Exception as e:
        logger.error("AI CSV preview failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to preview CSV with AI: {str(e)}"
        )

@router.post("/confirm_import")
async def confirm_import(
    trades: list = Body(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Confirm and save the previewed trades to database"""
    
    try:
        imported_count = 0
        skipped_count = 0
        errors = []
        
        for trade_data in trades:
            try:
                # Parse datetime strings to datetime objects
                filled_at_str = trade_data.get('filled_at')
                submitted_at_str = trade_data.get('submitted_at')
                
                filled_at_dt = None
                if filled_at_str:
                    try:
                        filled_at_dt = datetime.fromisoformat(filled_at_str.replace('Z', '+00:00'))
                    except Exception as e:
                        logger.warning("Failed to parse filled_at", value=filled_at_str, error=str(e))
                        filled_at_dt = datetime.utcnow()
                
                submitted_at_dt = None
                if submitted_at_str:
                    try:
                        submitted_at_dt = datetime.fromisoformat(submitted_at_str.replace('Z', '+00:00'))
                    except Exception as e:
                        logger.warning("Failed to parse submitted_at", value=submitted_at_str, error=str(e))
                
                # Check for duplicates based on symbol and timestamp
                existing_trade = db.query(Trade).filter(
                    Trade.user_id == current_user.id,
                    Trade.symbol == trade_data.get('symbol'),
                    Trade.filled_at == filled_at_dt
                ).first()
                
                if existing_trade:
                    skipped_count += 1
                    continue  # Skip duplicate
                
                # Create new trade from preview data with safe numeric values
                trade = Trade(
                    user_id=current_user.id,
                    account=trade_data.get('account', 'spot'),
                    venue=trade_data.get('venue', 'MANUAL'),
                    symbol=trade_data.get('symbol', ''),
                    side=trade_data.get('side', 'buy'),
                    qty=safe_numeric_value(trade_data.get('qty', 0)),
                    avg_price=safe_numeric_value(trade_data.get('avg_price', 0)),
                    fees=safe_numeric_value(trade_data.get('fees', 0)),
                    pnl=safe_numeric_value(trade_data.get('pnl', 0)),
                    filled_at=filled_at_dt,
                    submitted_at=submitted_at_dt,
                    raw=safe_raw_value(json.dumps(trade_data.get('raw', {})))
                )
                
                db.add(trade)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Failed to import trade: {str(e)}")
                logger.warning("Import error", error=str(e), trade_data=trade_data)
        
        db.commit()
        
        logger.info("CSV import confirmed", 
                   user_id=str(current_user.id), 
                   imported_count=imported_count,
                   skipped_count=skipped_count)
        
        message = f"Successfully imported {imported_count} trades"
        if skipped_count > 0:
            message += f" (skipped {skipped_count} duplicates)"
        
        return {
            "message": message,
            "imported_count": imported_count,
            "skipped_count": skipped_count,
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        logger.error("Failed to confirm import", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to save trades: {str(e)}"
        )

@router.post("/trades", response_model=TradeResponse)
async def create_trade(
    # Trade data as form fields
    venue: str = Form(...),
    symbol: str = Form(...),
    side: str = Form(...),
    qty: float = Form(...),
    avg_price: float = Form(...),
    fees: float = Form(0),
    pnl: float = Form(0),
    filled_at: str = Form(...),
    submitted_at: Optional[str] = Form(None),
    order_ref: Optional[str] = Form(None),
    session_id: Optional[str] = Form(None),
    account: Optional[str] = Form(None),
    raw: Optional[str] = Form(None),  # JSON string
    chart_image: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new trade manually with optional chart image"""
    
    try:
        # Parse raw data if provided
        raw_data = None
        if raw:
            try:
                raw_data = json.loads(raw)
            except json.JSONDecodeError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid JSON in raw field"
                )
        
        # Parse datetime strings to datetime objects
        try:
            filled_at_dt = datetime.fromisoformat(filled_at)
            submitted_at_dt = datetime.fromisoformat(submitted_at) if submitted_at else None
        except ValueError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid datetime format: {str(e)}"
            )
        
        # Handle chart image upload
        chart_image_path = None
        if chart_image:
            file_service = FileUploadService()
            chart_image_path = await file_service.upload_chart_image(chart_image, str(current_user.id))
        
        # Create trade record
        trade = Trade(
            user_id=current_user.id,
            venue=venue,
            symbol=symbol,
            side=side,
            qty=qty,
            avg_price=avg_price,
            fees=fees,
            pnl=pnl,
            filled_at=filled_at_dt,
            submitted_at=submitted_at_dt,
            order_ref=order_ref,
            session_id=session_id,
            account=account,
            raw=json.dumps(raw_data) if raw_data else None,
            chart_image=chart_image_path
        )
        
        db.add(trade)
        db.commit()
        db.refresh(trade)
        
        logger.info("Trade created", user_id=str(current_user.id), trade_id=str(trade.id))
        
        # Convert trade to response format, handling NaN/Infinity values
        try:
            trade_dict = {
                'id': trade.id,
                'user_id': trade.user_id,
                'account': trade.account,
                'venue': trade.venue,
                'symbol': trade.symbol,
                'side': trade.side,
                'qty': safe_numeric_value(trade.qty),
                'avg_price': safe_numeric_value(trade.avg_price),
                'fees': safe_numeric_value(trade.fees),
                'pnl': safe_numeric_value(trade.pnl),
                'submitted_at': trade.submitted_at,
                'filled_at': trade.filled_at,
                'order_ref': trade.order_ref,
                'session_id': trade.session_id,
                'raw': safe_raw_value(trade.raw),
                'chart_image': trade.chart_image
            }
            return TradeResponse(**trade_dict)
        except Exception as e:
            logger.error("Failed to convert trade to response", trade_id=trade.id, error=str(e))
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to process trade data"
            )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Failed to create trade", error=str(e), user_id=str(current_user.id))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create trade"
        )

@router.get("/trades")
async def get_trades(
    limit: Optional[int] = None,
    symbol: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get user's trades
    
    Args:
        limit: Optional limit on number of trades to return. If not specified, returns all trades.
        symbol: Optional filter by symbol
        start_date: Optional filter by start date (ISO format)
        end_date: Optional filter by end date (ISO format)
    """
    logger.info("=== STARTING GET_TRADES ENDPOINT ===")
    logger.info(f"User ID: {current_user.id}, Limit: {limit}, Symbol: {symbol}, Date Range: {start_date} to {end_date}")
    
    query = db.query(Trade).filter(Trade.user_id == current_user.id)
    
    if symbol:
        query = query.filter(Trade.symbol == symbol)
    
    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            query = query.filter(Trade.filled_at >= start_dt)
        except Exception as e:
            logger.warning("Invalid start_date", error=str(e))
    
    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query = query.filter(Trade.filled_at <= end_dt)
        except Exception as e:
            logger.warning("Invalid end_date", error=str(e))
    
    query = query.order_by(Trade.filled_at.desc())
    
    # Apply limit only if specified
    if limit:
        query = query.limit(limit)
    
    trades = query.all()
    logger.info(f"Fetched {len(trades)} trades from database")
    
    # Filter out corrupted trades and convert to response format
    trade_responses = []
    corrupted_count = 0
    
    for i, trade in enumerate(trades):
        try:
            # Check if trade is corrupted before processing
            is_corrupted, corruption_reason = is_trade_corrupted(trade)
            if is_corrupted:
                corrupted_count += 1
                logger.warning(f"Skipping corrupted trade {i+1}/{len(trades)}", 
                             trade_id=trade.id, reason=corruption_reason)
                continue
            
            logger.info(f"Processing trade {i+1}/{len(trades)}", trade_id=trade.id, 
                       qty=trade.qty, avg_price=trade.avg_price, fees=trade.fees, pnl=trade.pnl,
                       qty_type=type(trade.qty), avg_price_type=type(trade.avg_price))
            
            # Create a dict with safe numeric values
            trade_dict = {
                'id': trade.id,
                'user_id': trade.user_id,
                'account': trade.account,
                'venue': trade.venue,
                'symbol': trade.symbol,
                'side': trade.side,
                'qty': safe_numeric_value(trade.qty),
                'avg_price': safe_numeric_value(trade.avg_price),
                'fees': safe_numeric_value(trade.fees),
                'pnl': safe_numeric_value(trade.pnl),
                'submitted_at': trade.submitted_at,
                'filled_at': trade.filled_at,
                'order_ref': trade.order_ref,
                'session_id': trade.session_id,
                'raw': safe_raw_value(trade.raw),
                'chart_image': trade.chart_image
            }
            
            logger.info(f"Created trade_dict for trade {i+1}", trade_dict=trade_dict)
            trade_response = TradeResponse(**trade_dict)
            trade_responses.append(trade_response)
            logger.info(f"Successfully converted trade {i+1}")
            
        except Exception as e:
            corrupted_count += 1
            logger.error("Failed to convert trade to response", trade_id=trade.id, error=str(e), 
                        trade_data={
                            'qty': trade.qty, 'avg_price': trade.avg_price, 'fees': trade.fees, 'pnl': trade.pnl,
                            'qty_type': type(trade.qty), 'avg_price_type': type(trade.avg_price)
                        })
            continue
    
    logger.info(f"Successfully processed {len(trade_responses)} trades out of {len(trades)}")
    if corrupted_count > 0:
        logger.warning(f"Skipped {corrupted_count} corrupted trades to prevent JSON serialization errors")
    logger.info("About to return trade_responses", count=len(trade_responses))
    
    # Convert to dicts to ensure JSON serialization works
    result = []
    for trade_response in trade_responses:
        try:
            trade_dict = trade_response.model_dump()
            result.append(trade_dict)
        except Exception as e:
            logger.error("Failed to convert trade_response to dict", error=str(e))
            continue
    
    logger.info(f"Returning {len(result)} trades as dicts")
    return result

@router.get("/trades/{trade_id}", response_model=TradeResponse)
async def get_trade(
    trade_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get a single trade by ID"""
    
    trade = db.query(Trade).filter(
        Trade.id == trade_id,
        Trade.user_id == current_user.id
    ).first()
    
    if not trade:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Trade not found"
        )
    
    # Convert trade to response format, handling NaN/Infinity values
    try:
        trade_dict = {
            'id': trade.id,
            'user_id': trade.user_id,
            'account': trade.account,
            'venue': trade.venue,
            'symbol': trade.symbol,
            'side': trade.side,
            'qty': trade.qty if trade.qty is not None and not (isinstance(trade.qty, float) and (math.isnan(trade.qty) or math.isinf(trade.qty))) else None,
            'avg_price': trade.avg_price if trade.avg_price is not None and not (isinstance(trade.avg_price, float) and (math.isnan(trade.avg_price) or math.isinf(trade.avg_price))) else None,
            'fees': trade.fees if trade.fees is not None and not (isinstance(trade.fees, float) and (math.isnan(trade.fees) or math.isinf(trade.fees))) else None,
            'pnl': trade.pnl if trade.pnl is not None and not (isinstance(trade.pnl, float) and (math.isnan(trade.pnl) or math.isinf(trade.pnl))) else None,
            'submitted_at': trade.submitted_at,
            'filled_at': trade.filled_at,
            'order_ref': trade.order_ref,
            'session_id': trade.session_id,
            'raw': trade.raw,
            'chart_image': trade.chart_image
        }
        return TradeResponse(**trade_dict)
    except Exception as e:
        logger.error("Failed to convert trade to response", trade_id=trade.id, error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to process trade data"
        )

@router.put("/trades/{trade_id}", response_model=TradeResponse)
async def update_trade(
    trade_id: str,
    # Trade data as form fields
    venue: str = Form(...),
    symbol: str = Form(...),
    side: str = Form(...),
    qty: float = Form(...),
    avg_price: float = Form(...),
    fees: float = Form(0),
    pnl: float = Form(0),
    filled_at: str = Form(...),
    submitted_at: Optional[str] = Form(None),
    order_ref: Optional[str] = Form(None),
    session_id: Optional[str] = Form(None),
    account: Optional[str] = Form(None),
    raw: Optional[str] = Form(None),  # JSON string
    chart_image: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update a trade with optional chart image"""
    
    try:
        # Find trade
        trade = db.query(Trade).filter(
            Trade.id == trade_id,
            Trade.user_id == current_user.id
        ).first()
        
        if not trade:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Trade not found"
            )
        
        # Parse raw data if provided
        raw_data = None
        if raw:
            try:
                raw_data = json.loads(raw)
            except json.JSONDecodeError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid JSON in raw field"
                )
        
        # Parse datetime strings to datetime objects
        try:
            filled_at_dt = datetime.fromisoformat(filled_at)
            submitted_at_dt = datetime.fromisoformat(submitted_at) if submitted_at else None
        except ValueError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid datetime format: {str(e)}"
            )
        
        # Handle chart image upload
        chart_image_path = trade.chart_image  # Keep existing image by default
        if chart_image:
            file_service = FileUploadService()
            chart_image_path = await file_service.upload_chart_image(chart_image, str(current_user.id))
        
        # Update trade fields
        trade.venue = venue
        trade.symbol = symbol
        trade.side = side
        trade.qty = qty
        trade.avg_price = avg_price
        trade.fees = fees
        trade.pnl = pnl
        trade.filled_at = filled_at_dt
        trade.submitted_at = submitted_at_dt
        trade.order_ref = order_ref
        trade.session_id = session_id
        trade.account = account
        trade.raw = json.dumps(raw_data) if raw_data else trade.raw
        trade.chart_image = chart_image_path
        
        db.commit()
        db.refresh(trade)
        
        logger.info("Trade updated", user_id=str(current_user.id), trade_id=str(trade.id))
        
        # Convert trade to response format, handling NaN/Infinity values
        try:
            trade_dict = {
                'id': trade.id,
                'user_id': trade.user_id,
                'account': trade.account,
                'venue': trade.venue,
                'symbol': trade.symbol,
                'side': trade.side,
                'qty': safe_numeric_value(trade.qty),
                'avg_price': safe_numeric_value(trade.avg_price),
                'fees': safe_numeric_value(trade.fees),
                'pnl': safe_numeric_value(trade.pnl),
                'submitted_at': trade.submitted_at,
                'filled_at': trade.filled_at,
                'order_ref': trade.order_ref,
                'session_id': trade.session_id,
                'raw': safe_raw_value(trade.raw),
                'chart_image': trade.chart_image
            }
            return TradeResponse(**trade_dict)
        except Exception as e:
            logger.error("Failed to convert trade to response", trade_id=trade.id, error=str(e))
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to process trade data"
            )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Failed to update trade", error=str(e), user_id=str(current_user.id))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update trade"
        )

@router.delete("/trades/{trade_id}")
async def delete_trade(
    trade_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a trade"""
    
    # Find trade
    trade = db.query(Trade).filter(
        Trade.id == trade_id,
        Trade.user_id == current_user.id
    ).first()
    
    if not trade:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Trade not found"
        )
    
    # Delete associated journal entries first
    db.query(JournalEntry).filter(JournalEntry.trade_id == trade_id).delete()
    
    # Delete trade
    db.delete(trade)
    db.commit()
    
    logger.info("Trade deleted", user_id=str(current_user.id), trade_id=str(trade_id))
    
    return {"message": "Trade deleted successfully"}

@router.post("/entry", response_model=JournalEntryResponse)
async def create_journal_entry(
    entry_data: JournalEntryCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a journal entry"""
    
    # Verify trade belongs to user if trade_id provided
    if entry_data.trade_id:
        trade = db.query(Trade).filter(
            Trade.id == entry_data.trade_id,
            Trade.user_id == current_user.id
        ).first()
        
        if not trade:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Trade not found"
            )
    
    # Create journal entry
    journal_entry = JournalEntry(
        user_id=current_user.id,
        **entry_data.dict()
    )
    
    db.add(journal_entry)
    db.commit()
    db.refresh(journal_entry)
    
    logger.info("Journal entry created", user_id=str(current_user.id), entry_id=str(journal_entry.id))
    
    return JournalEntryResponse.from_orm(journal_entry)

@router.get("/entries", response_model=List[JournalEntryResponse])
async def get_journal_entries(
    trade_id: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get journal entries"""
    
    query = db.query(JournalEntry).filter(JournalEntry.user_id == current_user.id)
    
    if trade_id:
        query = query.filter(JournalEntry.trade_id == trade_id)
    
    entries = query.order_by(JournalEntry.ts.desc()).limit(limit).all()
    
    return [JournalEntryResponse.from_orm(entry) for entry in entries]

@router.get("/analytics")
async def get_analytics(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get trading analytics/KPIs"""
    
    # Get all trades for user
    trades = db.query(Trade).filter(Trade.user_id == current_user.id).all()
    
    if not trades:
        return {
            "total_trades": 0,
            "win_rate": 0,
            "total_pnl": 0,
            "avg_r": 0,
            "max_drawdown": 0,
            "best_symbol": None,
            "worst_symbol": None
        }
    
    # Calculate KPIs
    total_trades = len(trades)
    winning_trades = [t for t in trades if t.pnl > 0]
    win_rate = len(winning_trades) / total_trades if total_trades > 0 else 0
    
    total_pnl = sum(t.pnl for t in trades)
    
    # Calculate average R (risk/reward ratio)
    # This is simplified - in reality you'd need stop loss data
    avg_r = 0  # TODO: Implement proper R calculation
    
    # Calculate max drawdown
    cumulative_pnl = 0
    peak = 0
    max_dd = 0
    
    for trade in sorted(trades, key=lambda x: x.filled_at):
        cumulative_pnl += trade.pnl
        if cumulative_pnl > peak:
            peak = cumulative_pnl
        drawdown = peak - cumulative_pnl
        if drawdown > max_dd:
            max_dd = drawdown
    
    # Best/worst symbols
    symbol_pnl = {}
    for trade in trades:
        if trade.symbol not in symbol_pnl:
            symbol_pnl[trade.symbol] = 0
        symbol_pnl[trade.symbol] += trade.pnl
    
    best_symbol = max(symbol_pnl.items(), key=lambda x: x[1])[0] if symbol_pnl else None
    worst_symbol = min(symbol_pnl.items(), key=lambda x: x[1])[0] if symbol_pnl else None
    
    return {
        "total_trades": total_trades,
        "win_rate": round(win_rate * 100, 2),
        "total_pnl": float(total_pnl),
        "avg_r": avg_r,
        "max_drawdown": float(max_dd),
        "best_symbol": best_symbol,
        "worst_symbol": worst_symbol,
        "symbol_performance": symbol_pnl
    }

@router.post("/ingest_csv")
async def ingest_csv(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Import trades from CSV file"""
    
    try:
        # Validate file type
        if not file.filename.endswith('.csv'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only CSV files are supported"
            )
        
        # Read CSV content
        content = await file.read()
        df = pd.read_csv(io.StringIO(content.decode('utf-8')))
        
        # Parse trades using CSV parser
        csv_parser = CSVParser()
        parsed_trades = csv_parser.parse_trades(df)
        
        if not parsed_trades:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No valid trades found in CSV file"
            )
        
        # Import trades to database
        imported_count = 0
        errors = []
        
        for trade_data in parsed_trades:
            try:
                # Check for duplicates (by order_ref and filled_at)
                existing_trade = db.query(Trade).filter(
                    Trade.user_id == current_user.id,
                    Trade.order_ref == trade_data.get('order_ref'),
                    Trade.filled_at == trade_data.get('filled_at')
                ).first()
                
                if existing_trade:
                    continue  # Skip duplicate
                
                # Create new trade
                trade = Trade(
                    user_id=current_user.id,
                    account=trade_data.get('account', 'spot'),
                    venue=trade_data.get('venue', 'UNKNOWN'),
                    symbol=trade_data.get('symbol', ''),
                    side=trade_data.get('side', 'buy'),
                    qty=trade_data.get('qty', 0),
                    avg_price=trade_data.get('avg_price', 0),
                    fees=trade_data.get('fees', 0),
                    pnl=trade_data.get('pnl', 0),
                    filled_at=trade_data.get('filled_at'),
                    order_ref=trade_data.get('order_ref'),
                    raw=str(trade_data.get('raw', {}))
                )
                
                db.add(trade)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Failed to import trade: {str(e)}")
                logger.warning("CSV import error", error=str(e), trade_data=trade_data)
        
        db.commit()
        
        logger.info("CSV import completed", 
                   user_id=str(current_user.id), 
                   imported_count=imported_count,
                   total_trades=len(parsed_trades))
        
        return {
            "message": f"Successfully imported {imported_count} trades",
            "imported_count": imported_count,
            "total_trades": len(parsed_trades),
            "errors": errors[:10] if errors else []  # Limit error messages
        }
        
    except Exception as e:
        logger.error("CSV import failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to import CSV: {str(e)}"
        )
